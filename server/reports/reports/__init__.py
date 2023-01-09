import os
from db import Db
from queries import select_all_tasks
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from reports.models import Partida, Usuario
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

engine = create_engine("sqlite:///../database.sqlite", echo=True)
Session = sessionmaker(bind=engine)


def relative_path(path: str):
    absolute_path = os.path.dirname(__file__)
    relative_path = path
    full_path = os.path.join(absolute_path, relative_path)
    return full_path


def main():
    database = Db(relative_path("../../database.sqlite"))
    connection = database.connection

    wb = load_workbook(filename=relative_path("../reports.xlsx"))
    sheet_ranges = wb['kardex']
    sheet_ranges['A1'].value = 'Edwar martinez vale'
    # sheet_ranges.move_range('A13:L13', rows=8, cols=0)
    sheet_ranges.insert_rows(12, 10)
    thin = Side(border_style="thin", color="000000")
    for row in sheet_ranges.iter_rows(min_row=12, max_col=12, max_row=22):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.value = 'Edwar'
    # sheet_ranges['A12'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wb.save('test.xlsx')
    # print(sheet_ranges['A1'].value)

    # if connection:
    # print("1. Query task by priority:")
    # select_task_by_priority(conn, 1)
    # print("2. Query all tasks")
    # select_all_tasks(connection)

    # session = Session()
    # for x in session.query(Usuario):
    #     print(x.nombre, x.apellido)


if __name__ == '__main__':
    main()
